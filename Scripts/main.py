import shutil
import pandas as pd
from datetime import datetime
from datetime import timedelta
from datetime import time
import sqlalchemy
from sqlalchemy import engine, text, create_engine
import oracledb
import logging
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import config
import os


def convertir_hora(col):
    # intenta convertir directamente
    dt = pd.to_datetime(col, errors="coerce")

    # si ya parece número (Excel), convertir diferente
    if pd.api.types.is_numeric_dtype(col):
        dt = pd.to_datetime(col, unit="d", origin="1899-12-30", errors="coerce")

    return dt


#Rutas archivos de huelleros
#Rutas archivos de huelleros
r_caseta = config.r_caseta()
r_admin = config.r_admin()
r_dit = config.r_dit()

# preparar rutas temporales para copiar los archivos de huelleros
temp_caseta = os.path.join("temp", "temp_" + os.path.basename(r_caseta))
temp_admin = os.path.join("temp", "temp_" + os.path.basename(r_admin))
temp_dit = os.path.join("temp", "temp_" + os.path.basename(r_dit))

shutil.copy(r_caseta, temp_caseta)
shutil.copy(r_admin, temp_admin)
shutil.copy(r_dit, temp_dit)
df_CASETA = pd.read_excel(temp_caseta)
df_ADMIN = pd.read_excel(temp_admin)
df_DIT = pd.read_excel(temp_dit)
print(f"cabecera caseta: {df_CASETA.head()}")
print(f"cabecera admin: {df_ADMIN.head()}")
print(f"cabecera dit: {df_DIT.head()}")
os.remove(temp_caseta)
os.remove(temp_admin)
os.remove(temp_dit)

df_CASETA["HUELLERO"] = "H_CASETA_1" 
df_ADMIN["HUELLERO"] = "H_ADMIN"
df_DIT["HUELLERO"] = "H_DIT"

#fecha de ultimo procesamiento de los archivos
if os.path.exists(config.ultimo_proceso):
    with open(config.ultimo_proceso) as f:
        ult_proceso = datetime.strptime(f.readline(), "%Y-%m-%d %H:%M:%S")
else:
    ult_proceso = datetime.now()


# Fecha actual
hoy = datetime.now().date()

# Fecha inicial (3 días antes del último proceso)
fecha_inicio = (ult_proceso - timedelta(days=3))

df_huelleros = pd.concat([df_CASETA, df_DIT, df_ADMIN], ignore_index=True)
df_huelleros["FECHA Y HORA"] = pd.to_datetime(df_huelleros["FECHA Y HORA"])
df_huelleros = df_huelleros[df_huelleros["FECHA Y HORA"] >= fecha_inicio]

df_huelleros = df_huelleros.sort_values(by=["FECHA Y HORA"])

df_huelleros["FECHA"] = pd.to_datetime(df_huelleros["FECHA Y HORA"]).dt.date
df_huelleros["HORA"] = pd.to_datetime(df_huelleros["FECHA Y HORA"]).dt.time

#conectar a la base de datos
oracledb.init_oracle_client(lib_dir=r""+config.r_instantClient)
connection_url = f"oracle+oracledb://{config.USER}:{config.PASSWORD}@{config.HOST}:{config.PORT}/?service_name={config.SID}"

engine = engine = create_engine(
    connection_url,
    pool_pre_ping=True, 
    pool_recycle=1200 )

#consulta empleados activos

with open(config.r_query, "r") as file:
    query = file.read()

try:
    with engine.connect() as connection:
        df_emp = pd.DataFrame(connection.execute(text(query)))
except Exception as e:
    print (f"error en la base de datos: {e}")
engine.dispose()
df_emp = df_emp.rename(columns={'nit': 'NIT'})

#Cruce de todos los nits con todas las fechas
df_fechas = df_huelleros[["FECHA"]].drop_duplicates()
df_NITs_fechas = df_emp[["NIT"]].merge(df_fechas, how="cross")


#preparacion de dataframes para cruzar
df_huelleros = df_huelleros.rename(columns={'Documento': 'NIT'})
df_huelleros["NIT"] = df_huelleros["NIT"].astype(str)
df_NITs_fechas["NIT"] = df_NITs_fechas["NIT"].astype(str)

#merge de dataframes de fechas, empleados y registros de huelleros
df_all = pd.merge(df_NITs_fechas, df_emp, on="NIT")
df_all = pd.merge(df_all, df_huelleros, on=["NIT", "FECHA"], how="left")

#hora y huellero de ingreso y salida
df_ingreso = df_huelleros.sort_values("FECHA Y HORA") \
    .groupby(["NIT","FECHA"], as_index=False) \
    .first()[["NIT","FECHA","FECHA Y HORA","HUELLERO"]]


df_salida = df_huelleros.sort_values("FECHA Y HORA") \
    .groupby(["NIT","FECHA"], as_index=False) \
    .last()[["NIT","FECHA","FECHA Y HORA","HUELLERO"]]


df_ingreso = df_ingreso.rename(columns={
    "FECHA Y HORA": "HORA_INGRESO",
    "HUELLERO": "H_INGRESO"
})

df_salida = df_salida.rename(columns={
    "FECHA Y HORA": "HORA_SALIDA",
    "HUELLERO": "H_SALIDA"
})

df_all = df_all.drop_duplicates(subset=["NIT", "FECHA"])
df_all = pd.merge(df_all,df_ingreso,on =["NIT","FECHA"], how="left")
df_all = pd.merge(df_all,df_salida,on =["NIT","FECHA"], how="left")

#limpieza de columnas y filas innecesarias
df_all = df_all.drop(columns=["FECHA Y HORA", "HORA", "HUELLERO"])
df_all = df_all.drop_duplicates()

#reafirmacipon de formatos
df_all["HORA_INGRESO"] = pd.to_datetime(df_all["HORA_INGRESO"], errors="coerce")
df_all["HORA_SALIDA"]  = pd.to_datetime(df_all["HORA_SALIDA"], errors="coerce")

#horas laboradas
df_all["HORAS_LABORADAS"] = df_all["HORA_SALIDA"] - df_all["HORA_INGRESO"]
df_all["CANTIDAD_HORAS_LABORADAS"] = (df_all["HORAS_LABORADAS"].dt.total_seconds()/3600).__round__(2)

##############limpieza y formato de datos(datos nulos, formatos, etc.)############################
#solo salida
df_all.loc[
    (df_all["HORA_INGRESO"] == df_all["HORA_SALIDA"]) & (df_all["HORA_INGRESO"].dt.hour > 12),
    ["HORA_INGRESO", "H_INGRESO"]
] = None
#solo ingreso
df_all.loc[
    (df_all["HORA_INGRESO"] == df_all["HORA_SALIDA"]) & (df_all["HORA_SALIDA"].dt.hour < 12),
    ["HORA_SALIDA", "H_SALIDA"]
] = None

#datos nulos
df_all.loc[
    df_all["HORA_INGRESO"].isna(),
    ["H_INGRESO"]
] = "NO REGISTRA"
df_all.loc[
    df_all["HORA_SALIDA"].isna(),
    ["H_SALIDA"]
] = "NO REGISTRA"
df_all.loc[
    df_all["CANTIDAD_HORAS_LABORADAS"].isna(),
    ["CANTIDAD_HORAS_LABORADAS"]
] = 0

df_all["HORAS_LABORADAS"] = df_all["HORAS_LABORADAS"].fillna(pd.Timedelta(0))

#reorganizar y renombrar columnas
df_all = df_all.reindex(columns = ["NIT","empresa","Nombre", "Cargo", "Dependencia", "FECHA", "HORA_INGRESO", "HORA_SALIDA", "HORAS_LABORADAS","CANTIDAD_HORAS_LABORADAS","H_INGRESO", "H_SALIDA"])
df_all = df_all.rename(columns={
    "NIT": "DOCUMENTO",
    "empresa": "EMPRESA",
    "Nombre" : "NOMBRE",
    "Cargo" : "NOMBRE_CARGO",
    "Dependencia" : "NOMBRE_DEPENDENCIA",
    })

#eliminar duplicados
df_all = df_all.drop_duplicates()


# Duración (horas trabajadas)
df_all["HORAS_LABORADAS"] = pd.to_timedelta(df_all["HORAS_LABORADAS"], errors='coerce')
df_all["HORAS_LABORADAS"] = df_all["HORAS_LABORADAS"].dt.total_seconds() / 86400


                  
#Cargue de datos al excel
if os.path.isfile(config.r_historico):
    df_existente = pd.read_excel(config.r_historico)
    # Normalizar tipos para evitar falsos duplicados
    df_all["DOCUMENTO"] = df_all["DOCUMENTO"].astype(str)
    df_existente["DOCUMENTO"] = df_existente["DOCUMENTO"].astype(str)

    df_all["FECHA"] = pd.to_datetime(df_all["FECHA"]).dt.date
    df_existente["FECHA"] = pd.to_datetime(df_existente["FECHA"]).dt.date

    # Concatenar y eliminar duplicados
    df_total = pd.concat([df_existente, df_all], ignore_index=True)
    
    df_total = df_total.drop_duplicates(subset=[
        "DOCUMENTO", "FECHA"])
    df_total = df_total.sort_values(by = "FECHA", ascending= True, ignore_index= True)

    # Convertir solo si NO es numérico    
    horas_ingreso = convertir_hora(df_total["HORA_INGRESO"])
    horas_salida  = convertir_hora(df_total["HORA_SALIDA"])
    df_total["HORA_INGRESO"] = horas_ingreso.dt.time
    df_total["HORA_SALIDA"] = horas_salida.dt.time

    
# Sobrescribir archivo limpio
    with pd.ExcelWriter(config.r_historico, engine="openpyxl") as writer:
        df_total.to_excel(writer, index=False)

        ws = writer.sheets["Sheet1"]

        for row in range(2, len(df_total) + 2):
            ws[f"G{row}"].number_format = "hh:mm"      # ingreso
            ws[f"H{row}"].number_format = "hh:mm"      # salida
            ws[f"I{row}"].number_format = "[hh]:mm"    # duración
        print("archivo actualizado")
else:

    with pd.ExcelWriter(config.r_historico, engine="openpyxl") as writer:
        df_all.to_excel(writer, index=False)

        ws = writer.sheets["Sheet1"]

        for row in range(2, len(df_all) + 2):
            ws[f"G{row}"].number_format = "hh:mm"      # ingreso
            ws[f"H{row}"].number_format = "hh:mm"      # salida
            ws[f"I{row}"].number_format = "[hh]:mm"    # duración
        print("archivo creado")

with open("ultimo_proceso.txt", mode= "w") as f:
        f.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

